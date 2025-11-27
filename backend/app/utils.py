"""Utility functions for file processing."""

from __future__ import annotations

import json
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from werkzeug.datastructures import FileStorage

from .config import Config

# ====== 현재 월_년 형식 생성 (예: "_11_25") ======
def get_current_month_year_suffix() -> str:
    """오늘 날짜 기준으로 '_월_년' 형식의 접미사 반환 (예: '_11_25')"""
    now = datetime.now()
    return f"_{now.month:02d}_{now.strftime('%y')}"

# ====== 파일별 타깃 코드 설정 (suffix 제외한 기본 파일명 기준) ======
FILE_CODE_MAP = {
    "SIN_Timeseries_BCI 5TC": ["534544"],
    "SIN_Timeseries_BHSI": ["542370"],
    "SIN_Timeseries_BPI": ["542181"],
    "SIN_Timeseries_BSI": ["552787"],
    "SIN_Timeseries_Cape Fleet Development": ["30201"],
    "SIN_Timeseries_Dry Bulk Trade": ["534399", "534400", "534401", "534409"],
    "SIN_Timeseries_Handy Fleet Development": ["30209"],
    "SIN_Timeseries_Panamax Fleet Development": ["30204"],
    "SIN_Timeseries_Supramax Fleet Development": ["30207"],
    "SIN_Timeseries_Containership Fleet Development Annual": ["30977", "534501"],
    "SIN_Timeseries_Container Trade Annual": ["548622"],
    "SIN_Timeseries_Container SCFI": ["534015", "534018", "534021"],
}

# Allowed file extensions
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
TEXT_EXTENSIONS = {".md", ".html", ".htm", ".json"}
PDF_EXTENSIONS = {".pdf"}


def is_allowed_excel_file(filename: str) -> bool:
    """Check if file has an allowed Excel extension."""
    return Path(filename).suffix.lower() in EXCEL_EXTENSIONS


def is_allowed_text_file(filename: str) -> bool:
    """Check if file has an allowed text extension."""
    return Path(filename).suffix.lower() in TEXT_EXTENSIONS


def is_allowed_pdf_file(filename: str) -> bool:
    """Check if file has an allowed PDF extension."""
    return Path(filename).suffix.lower() in PDF_EXTENSIONS


def generate_unique_filename(original_filename: str) -> str:
    """Generate a unique filename to prevent collisions."""
    ext = Path(original_filename).suffix
    unique_id = uuid.uuid4().hex
    return f"{unique_id}{ext}"


def save_upload_file(file: FileStorage, folder: Path) -> Path:
    """Save uploaded file to the specified folder and return the saved path."""
    if not folder.exists():
        folder.mkdir(parents=True, exist_ok=True)

    unique_filename = generate_unique_filename(file.filename or "unknown")
    save_path = folder / unique_filename
    file.save(str(save_path))
    return save_path


def find_date_column(columns: list[str]) -> str | None:
    """
    컬럼명 리스트에서 'date' 가 들어간 첫 번째 컬럼을 반환.
    없으면 None.
    """
    for col in columns:
        if "date" in str(col).strip().lower():
            return col
    return None


def select_columns_by_codes(columns: list[str], codes: list[str]) -> list[str]:
    """
    코드가 포함된 컬럼만 선택하여 반환.
    
    Args:
        columns: 전체 컬럼명 리스트
        codes: 찾을 코드 리스트
    
    Returns:
        선택된 컬럼명 리스트 (Date 컬럼 + 코드가 포함된 컬럼들)
    """
    selected_cols = []
    
    # Date 컬럼 찾기
    date_col = find_date_column(columns)
    if date_col:
        selected_cols.append(date_col)
    
    # 코드가 들어있는 컬럼 찾기
    for code in codes:
        matched = [c for c in columns if code in str(c)]
        selected_cols.extend(matched)
    
    # 중복 제거 (순서 유지)
    selected_cols = list(dict.fromkeys(selected_cols))
    
    return selected_cols


def build_headers_from_rows(ws, row_top: int = 4, row_mid: int = 5, row_bottom: int = 6) -> list[str]:
    """
    주어진 워크시트에서 row_top, row_mid, row_bottom 행을 읽어
    '위 행값 - 중간 행값 - 아래 행값' 형태로 컬럼명 리스트를 생성.
    모든 컬럼을 읽기 위해 max_column까지 처리.
    """
    # 최대 컬럼 수 확인 (데이터가 있는 마지막 컬럼까지)
    max_col = ws.max_column
    
    # 각 행의 모든 셀 값 읽기
    row_top_vals = []
    row_mid_vals = []
    row_bottom_vals = []
    
    for col_idx in range(1, max_col + 1):
        cell_top = ws.cell(row=row_top, column=col_idx)
        cell_mid = ws.cell(row=row_mid, column=col_idx)
        cell_bottom = ws.cell(row=row_bottom, column=col_idx)
        row_top_vals.append(cell_top.value)
        row_mid_vals.append(cell_mid.value)
        row_bottom_vals.append(cell_bottom.value)
    
    headers = []
    
    for v4, v5, v6 in zip(row_top_vals, row_mid_vals, row_bottom_vals):
        t4 = (str(v4).strip() if v4 is not None else "")
        t5 = (str(v5).strip() if v5 is not None else "")
        t6 = (str(v6).strip() if v6 is not None else "")
        
        # 세 행의 값들을 합치기
        parts = [t4, t5, t6]
        parts = [p for p in parts if p]  # 빈 문자열 제거
        
        if len(parts) > 0:
            name = " - ".join(parts)
        else:
            name = ""
        
        headers.append(name)
    
    return headers


def extract_header_labels(excel_path: Path, sheet_name: str | None = None, row_top: int = 4, row_mid: int = 5, row_bottom: int = 6) -> tuple[list[str], list[str], list[str], str]:
    """
    엑셀 시트의 4행, 5행, 6행을 합쳐서 컬럼명을 생성하는 함수
    
    Args:
        excel_path: 엑셀 파일 경로
        sheet_name: 시트 이름 (None이면 첫 번째 시트)
        row_top: 상단 행 번호 (기본값: 4)
        row_mid: 중간 행 번호 (기본값: 5)
        row_bottom: 하단 행 번호 (기본값: 6)
    
    Returns:
        tuple: (labeled_cols, ordered_cargos, units, upper_bigtext)
    """
    from openpyxl import load_workbook
    import pandas as pd
    
    # ---------- 1) 워크북/시트 로드 ----------
    wb = load_workbook(excel_path, data_only=True)
    
    # 시트 이름이 지정되지 않았으면 첫 번째 시트 사용
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}'를 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}")
    
    ws = wb[sheet_name]
    
    # ---------- 2) 4행+5행+6행에서 컬럼명 생성 ----------
    labeled_cols = build_headers_from_rows(ws, row_top=row_top, row_mid=row_mid, row_bottom=row_bottom)
    
    # ---------- 3) 상단 텍스트 수집 (참고용) ----------
    upper_text = []
    for row in ws.iter_rows(min_row=1, max_row=max(row_top, row_mid, row_bottom) + 1):
        line = " ".join([str(cell.value) for cell in row if cell.value])
        if line.strip():
            upper_text.append(line)
    
    upper_bigtext = "\n".join(upper_text)
    
    # ---------- 4) Cargo 명 패턴 추출 (참고용) ----------
    cargo_pattern = r"(Iron Ore|Coal|Grain|Minor Bulk|Other|Bauxite|Alumina|Nickel Ore|Steel|Fertilizer|Coke|Scrap|Total)"
    cargos = re.findall(cargo_pattern, upper_bigtext, flags=re.IGNORECASE)
    cargos = [c.title() for c in cargos]
    
    # 중복 제거하면서 순서 유지
    seen = set()
    ordered_cargos = []
    for c in cargos:
        if c not in seen:
            seen.add(c)
            ordered_cargos.append(c)
    
    # ---------- 5) 단위 추출 (참고용) ----------
    units = []
    for col_name in labeled_cols:
        text = str(col_name).lower()
        if "billion" in text:
            units.append("billion tonne-miles")
        elif "million" in text:
            units.append("million tonnes")
        elif "y/y" in text or "%" in text or "yoy" in text:
            units.append("yoy")
        else:
            units.append("unknown")
    
    return labeled_cols, ordered_cargos, units, upper_bigtext


def read_excel_file(file_path: Path, extract_labels: bool = True, sheet_name: str | None = None, original_filename: str | None = None) -> dict:
    """Read Excel file and return data as a dictionary."""
    import pandas as pd
    import zipfile

    # Verify file is actually an Excel file by checking file header
    try:
        with open(file_path, 'rb') as f:
            header = f.read(8)
            # XLSX files start with PK (ZIP signature)
            # XLS files start with D0 CF 11 E0 A1 B1 1A E1 (OLE2 signature)
            is_xlsx = header[:2] == b'PK'
            is_xls = header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
            
            if not (is_xlsx or is_xls):
                # Try to verify as ZIP (XLSX is a ZIP archive)
                try:
                    with zipfile.ZipFile(file_path, 'r') as zf:
                        if 'xl/workbook.xml' not in zf.namelist() and '[Content_Types].xml' not in zf.namelist():
                            raise ValueError(
                                f"파일이 올바른 엑셀 파일 형식이 아닙니다. "
                                f"파일 헤더를 확인한 결과 엑셀 파일 시그니처가 발견되지 않았습니다."
                            )
                except zipfile.BadZipFile:
                    if not is_xls:
                        raise ValueError(
                            f"파일이 올바른 엑셀 파일 형식이 아닙니다. "
                            f"파일이 손상되었거나 다른 형식의 파일일 수 있습니다."
                        )
    except Exception as e:
        # If header check fails, continue anyway - let pandas try
        pass

    # Determine file extension to choose appropriate engine
    ext = file_path.suffix.lower()
    
    # Try different engines based on file extension
    if ext == '.xlsx':
        engines = ['openpyxl', 'xlrd', None]
    elif ext == '.xls':
        engines = ['xlrd', 'openpyxl', None]
    else:
        engines = ['openpyxl', 'xlrd', None]
    
    df = None
    errors = []
    
    # 시트 이름이 지정되지 않았으면 첫 번째 시트 사용
    if sheet_name is None:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
            wb.close()
        except Exception:
            sheet_name = None
    
    # 7행부터 데이터 읽기 (skiprows=6: 0~5행까지 스킵, 6행부터 읽기 = 엑셀의 7행부터)
    skiprows = 6
    
    for engine in engines:
        try:
            if engine:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine, header=None, skiprows=skiprows)
            else:
                # Let pandas auto-detect
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=skiprows)
            break
        except Exception as e:
            error_msg = f"Engine '{engine or 'auto'}': {str(e)}"
            errors.append(error_msg)
            continue
    
    if df is None:
        error_details = "\n".join(errors)
        raise ValueError(
            f"엑셀 파일을 읽을 수 없습니다. 파일이 손상되었거나 지원되지 않는 형식일 수 있습니다.\n"
            f"시도한 엔진들:\n{error_details}\n"
            f"파일 경로: {file_path}\n"
            f"파일 확장자: {ext}"
        )
    
    # 컬럼명 추출 기능 활성화 시 (6행부터 읽기 전에 컬럼명 먼저 추출)
    labeled_cols = None
    cargos = None
    units = None
    upper_text = None
    if extract_labels and ext == '.xlsx':
        try:
            labeled_cols, cargos, units, upper_text = extract_header_labels(file_path, sheet_name=sheet_name)
        except Exception as e:
            # 추출 실패해도 계속 진행
            pass
    
    # 컬럼명 적용
    if labeled_cols is not None:
        # 실제 컬럼 개수만큼 header 잘라서 적용
        if len(labeled_cols) < df.shape[1]:
            # 컬럼 수가 더 많은 경우: 부족한 만큼 빈 문자열로 채우기
            labeled_cols = labeled_cols + [""] * (df.shape[1] - len(labeled_cols))
        elif len(labeled_cols) > df.shape[1]:
            # 헤더가 더 긴 경우: 앞부분만 사용
            labeled_cols = labeled_cols[:df.shape[1]]
        
        df.columns = labeled_cols
        columns = labeled_cols
    else:
        # 컬럼명 추출 실패 시 기본 컬럼명 사용
        columns = [f"Column{i+1}" for i in range(df.shape[1])]
        df.columns = columns
    
    # NaN, NaT, inf 값을 None으로 변환 (JSON 호환성을 위해)
    df_cleaned = df.replace([float('nan'), float('inf'), float('-inf')], None)
    df_cleaned = df_cleaned.where(pd.notnull(df_cleaned), None)
    
    # dict로 변환
    data = df_cleaned.to_dict(orient="records")
    
    # 추가로 NaN 값이 남아있을 수 있으므로 재귀적으로 처리
    def clean_nan_values(obj):
        """재귀적으로 NaN, inf 값을 None으로 변환"""
        if isinstance(obj, dict):
            return {k: clean_nan_values(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [clean_nan_values(item) for item in obj]
        elif isinstance(obj, float):
            if pd.isna(obj) or obj == float('inf') or obj == float('-inf'):
                return None
            return obj
        return obj
    
    data = clean_nan_values(data)
    
    # 파일명 기반 코드 매칭 및 컬럼 필터링
    selected_columns = columns
    filtered_data = data
    codes_used = []
    
    # 파일명에서 코드 찾기 (원본 파일명 우선 사용)
    filename_for_matching = original_filename if original_filename else file_path.name
    
    # suffix를 제거한 기본 파일명으로 매칭
    import re
    base_filename = re.sub(r'_\d{2}_\d{2}\.xlsx$', '', filename_for_matching, flags=re.IGNORECASE)
    
    if base_filename in FILE_CODE_MAP:
        codes = FILE_CODE_MAP[base_filename]
        codes_used = codes
        
        # 코드가 포함된 컬럼만 선택
        selected_columns = select_columns_by_codes(columns, codes)
        
        if selected_columns:
            # 선택된 컬럼만으로 데이터 필터링
            filtered_data = []
            for row in data:
                filtered_row = {col: row.get(col) for col in selected_columns if col in row}
                filtered_data.append(filtered_row)
        else:
            # 매칭되는 컬럼이 없으면 전체 컬럼 사용
            selected_columns = columns
            filtered_data = data
    
    result = {
        "columns": columns,  # 전체 컬럼명
        "selected_columns": selected_columns,  # 선택된 컬럼명 (코드 매칭 결과)
        "data": filtered_data,  # 필터링된 데이터
        "row_count": len(filtered_data),
        "column_count": len(selected_columns),
        "codes_used": codes_used,  # 사용된 코드 목록
    }
    
    # 컬럼명 추출 정보 추가 (이미 추출한 정보 재사용)
    if extract_labels and ext == '.xlsx' and labeled_cols is not None:
        try:
            # 이미 추출한 정보 재사용 (중복 호출 방지)
            result["labeled_columns"] = labeled_cols
            
            # 선택된 컬럼에 해당하는 라벨링된 컬럼명도 필터링
            if selected_columns != columns:
                # 원본 컬럼명과 라벨링된 컬럼명 매칭
                selected_labeled = []
                for orig_col in selected_columns:
                    if orig_col in columns:
                        idx = columns.index(orig_col)
                        if idx < len(labeled_cols):
                            selected_labeled.append(labeled_cols[idx])
                result["selected_labeled_columns"] = selected_labeled
            else:
                result["selected_labeled_columns"] = labeled_cols
            
            result["cargos"] = cargos
            result["units"] = units
            result["header_text"] = upper_text
        except Exception as e:
            # 추출 실패해도 기본 정보는 반환
            result["label_extraction_error"] = str(e)
    
    return result


def save_filtered_excel(data: list[dict], columns: list[str], output_path: Path) -> Path:
    """
    필터링된 데이터를 엑셀 파일로 저장.
    
    Args:
        data: 저장할 데이터 (딕셔너리 리스트)
        columns: 컬럼명 리스트
        output_path: 저장할 파일 경로
    
    Returns:
        저장된 파일 경로
    """
    import pandas as pd
    
    # DataFrame 생성
    df = pd.DataFrame(data)
    
    # 컬럼 순서 정렬
    if columns:
        # 존재하는 컬럼만 선택
        available_cols = [col for col in columns if col in df.columns]
        df = df[available_cols]
    
    # 엑셀 파일로 저장
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    
    return output_path


def filter_data_by_year(data: list[dict], date_column: str, year: int = 2020) -> list[dict]:
    """
    2020년 이후 데이터만 필터링.
    
    Args:
        data: 필터링할 데이터
        date_column: 날짜 컬럼명
        year: 기준 연도 (기본값: 2020)
    
    Returns:
        필터링된 데이터
    """
    from datetime import datetime
    
    filtered = []
    for row in data:
        date_value = row.get(date_column)
        if date_value is None:
            continue
        
        # 날짜 파싱 시도
        try:
            if isinstance(date_value, str):
                # 다양한 날짜 형식 시도
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
                    try:
                        dt = datetime.strptime(date_value, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    # 파싱 실패 시 스킵
                    continue
            elif isinstance(date_value, datetime):
                dt = date_value
            else:
                continue
            
            if dt.year >= year:
                filtered.append(row)
        except Exception:
            continue
    
    return filtered


def group_datasets_by_category() -> dict:
    """
    업로드된 데이터셋을 카테고리별로 그룹화.
    현재 월_년 접미사가 포함된 파일만 필터링.
    
    Returns:
        {
            "drybulk_trade": [dataset_ids],
            "fleet_development": [dataset_ids],
            "indices": [dataset_ids],
            "container_trade_fleet": [dataset_ids],
            "scfi_weekly": [dataset_ids]
        }
    """
    from .database import get_session
    from .models import Dataset
    
    session = get_session()
    try:
        datasets = session.query(Dataset).all()
        
        groups = {
            "drybulk_trade": [],
            "fleet_development": [],
            "indices": [],
            "container_trade_fleet": [],
            "scfi_weekly": []
        }
        
        # 현재 월_년 접미사 생성
        current_suffix = get_current_month_year_suffix().lower()
        
        for dataset in datasets:
            filename = dataset.original_filename.lower()
            
            # 현재 월_년 접미사가 포함된 파일만 처리
            if current_suffix not in filename:
                continue
            
            if "dry bulk trade" in filename:
                groups["drybulk_trade"].append(dataset.id)
            elif "fleet development" in filename:
                if "containership" in filename or ("container" in filename and "fleet" in filename):
                    groups["container_trade_fleet"].append(dataset.id)
                else:
                    groups["fleet_development"].append(dataset.id)
            elif any(x in filename for x in ["bci", "bhsi", "bpi", "bsi"]):
                groups["indices"].append(dataset.id)
            elif "container trade" in filename:
                groups["container_trade_fleet"].append(dataset.id)
            elif "container scfi" in filename:
                groups["scfi_weekly"].append(dataset.id)
        
        return groups
    finally:
        session.close()


def convert_text_to_html(file_path: Path, content_type: str) -> str:
    """Convert text file (Markdown or DOCX) to HTML."""
    if content_type == "markdown":
        import markdown

        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return markdown.markdown(content, extensions=["extra", "codehilite"])

    elif content_type == "docx":
        from docx import Document

        doc = Document(file_path)
        html_parts = ["<div>"]
        for para in doc.paragraphs:
            if para.text.strip():
                html_parts.append(f"<p>{para.text}</p>")
        html_parts.append("</div>")
        return "\n".join(html_parts)
    elif content_type == "html":
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    elif content_type == "json":
        import json

        with open(file_path, "r", encoding="utf-8") as f:
            try:
                payload = json.load(f)
                pretty = json.dumps(payload, ensure_ascii=False, indent=2)
            except json.JSONDecodeError:
                f.seek(0)
                pretty = f.read()
        return f"<pre><code>{pretty}</code></pre>"

    return "<p>吏?먰븯吏 ?딅뒗 ?뚯씪 ?뺤떇?낅땲??</p>"


def extract_pdf_info(file_path: Path, include_full_text: bool = False) -> dict:
    """Extract basic PDF metadata and optionally the entire text content."""

    try:
        import PyPDF2

        with open(file_path, "rb") as f:
            pdf_reader = PyPDF2.PdfReader(f)
            page_count = len(pdf_reader.pages)

            # File size in KB
            file_size_kb = file_path.stat().st_size / 1024

            # Preview text (first page, max 500 chars)
            preview_text = ""
            if page_count > 0:
                try:
                    first_page = pdf_reader.pages[0]
                    preview_text = (first_page.extract_text() or "")[:500]
                except Exception:
                    preview_text = "[텍스트 추출 불가]"

            full_text = None
            if include_full_text:
                extracted_pages: list[str] = []
                for page in pdf_reader.pages:
                    try:
                        extracted = page.extract_text() or ""
                    except Exception:
                        extracted = ""
                    if extracted.strip():
                        extracted_pages.append(extracted.strip())

                full_text = "\n\n".join(extracted_pages) if extracted_pages else "[텍스트 추출 불가]"

            return {
                "page_count": page_count,
                "file_size_kb": round(file_size_kb, 2),
                "preview_text": preview_text or "[텍스트 추출 불가]",
                "full_text": full_text if include_full_text else None,
            }
    except ImportError:
        file_size_kb = file_path.stat().st_size / 1024
        fallback = "[PDF 라이브러리가 설치되지 않았습니다]"
        return {
            "page_count": 0,
            "file_size_kb": round(file_size_kb, 2),
            "preview_text": fallback,
            "full_text": fallback if include_full_text else None,
        }
    except Exception as e:
        file_size_kb = file_path.stat().st_size / 1024
        fallback = f"[PDF 처리 오류: {str(e)}]"
        return {
            "page_count": 0,
            "file_size_kb": round(file_size_kb, 2),
            "preview_text": fallback,
            "full_text": fallback if include_full_text else None,
        }


